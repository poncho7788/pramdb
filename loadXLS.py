from openpyxl import load_workbook
from openpyxl.utils.cell import *



class AssessmentXLS:

    def __init__(self):
        self.book = load_workbook('assessment.xlsx')
        self.ws = self.book['Assessment']


        r=self.book.defined_names['AssetName']
        dests=list(r.destinations)[0][1]
        self.assetName=self.ws[dests].value

        r=self.book.defined_names['AssetType']
        dests=list(r.destinations)[0][1]
        self.assetType=self.ws[dests].value

        self.impacts = self.table_to_list('TblImpact')
        self.asls = self.table_to_list('TblAssessment')
        self.scenarios=self.table_to_list('TblScenarios')

    # Needs a table identifier
    # Returns a list of where each element corresponds to a row in the list, being a dictionary with the title of the columns as keys
    # if the first column is blank, skips the row
    def table_to_list(self, tableId):

        tab = self.ws.tables[tableId]
        titles = tab.column_names
        # range=self.book.defined_names['TblImpact'].value
        rango = self.ws.tables[tableId].ref
        pattern = '([A-Z]+)([0-9]+)\:([A-Z]+)([0-9]+)'
        res = re.match(pattern, rango)

        a = res.group(1)
        colIni = column_index_from_string(a)

        rowIni = int(res.group(2))
        rowIni += 1  # skip the title

        a = res.group(3)
        colEnd = column_index_from_string(a)

        rowEnd = int(res.group(4))

        records = []
        for row in range(rowIni, rowEnd + 1):
            item = {}
            cont = 0
            for t in titles:
                item[t] = self.ws.cell(row, colIni + cont).value
                cont += 1
            if item[titles[0]]:
                records.append(item)
        return records

    def range_char(self, start, stop):
        return (chr(n) for n in range(ord(start), ord(stop) + 1))

